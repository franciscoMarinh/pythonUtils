from openpyxl import Workbook, load_workbook

def validateNotIsNone(data):
    valite = False
    for i in data:
        if(i == None):
            valite = True
    return valite

def getColumValue(sheetBook, row, column):
    return sheetBook.cell(row=row, column=column).value

def readExcelRows(filePath):
    wb = load_workbook(filePath)
    firstSheet = wb.sheetnames[0]
    sheetBook = wb[firstSheet]
    data = []
    for i in range(2,400,1):
        arrayTemp = []
        arrayTemp.append(getColumValue(sheetBook, i,1))
        arrayTemp.append(getColumValue(sheetBook, i,2))
        arrayTemp.append(getColumValue(sheetBook, i,3))
        if(validateNotIsNone(arrayTemp)):
            break
        data.append(arrayTemp)
    return data